from openpyxl import load_workbook
from posto import Posto
from logger import Logger
from datetime import datetime
from config.config import PLANILHA_DIR


class Planilha():

    def __init__(self):
        # self.df = pd.DataFrame()
        self.arquivo = PLANILHA_DIR
        self.logger = Logger()

    def salva_planilha(self, lista_postos:Posto):
        workbook = load_workbook(self.arquivo)
        sheet = workbook['Precos']

        data_e_hora = datetime.now().strftime("%d/%m-%H:%M")
        sheet.cell(row=1, column=1).value = data_e_hora

        for posto in lista_postos:
            cif_values = [posto.cif_etanol, posto.cif_gasolina_ad, posto.cif_gasolina, posto.cif_s10, posto.cif_s500]
            fob_values = [posto.fob_etanol, posto.fob_gasolina_ad, posto.fob_gasolina, posto.fob_s10, posto.fob_s500]

            if posto.nome == "Ipiranga Pit Stop":
                columns = [2, 3]
            elif posto.nome == "Ipiranga Distrito":
                columns = [4, 5]
            elif posto.nome == "Ipiranga Gas Station":
                columns = [6, 7]
            elif posto.nome == "Ipiranga Itirapuã":
                columns = [8, 9]
            elif posto.nome == "Ipiranga PPP":
                columns = [10, 11]
            elif posto.nome == "Vibra Janjão":
                columns = [12, 13]
            elif posto.nome == "Ipiranga TRR 1":
                columns = [14, 15]
            elif posto.nome == "Ipiranga TRR 2":
                columns = [16, 17]
            elif posto.nome == "Vibra TRR":
                columns = [18, 19]
            elif posto.nome == "Raízen TRR":
                columns = [20, 21]

            row = 3
            for cif, fob in zip(cif_values, fob_values):
                sheet.cell(row=row, column=columns[0]).value = cif
                sheet.cell(row=row, column=columns[1]).value = fob
                row += 1

        workbook.save(self.arquivo)
        self.logger.log("Preços salvos na planilha.")
