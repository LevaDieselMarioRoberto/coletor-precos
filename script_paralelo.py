import ipiranga
import vibra
import raizen
import ipiranga_postos
import vibra_janjao
from funcoes import imprime_precos, cria_dicionario_para_df

import threading
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from time import time

global precos_ipr1, precos_ipr2, precos_vbr, precos_rzn, tempo_ex_ipr, tempo_ex_vbr, tempo_ex_rzn
global precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop, tempo_ex_ipr_postos
global precos_vbr_janjao, tempo_ex_vbr_janjao
global ipiranga_deu_erro, vibra_deu_erro, raizen_deu_erro, ipiranga_postos_deu_erro, vibra_janjao_deu_erro


def coleta_e_imprime(preco_func, nome_posto):

    global precos_ipr1, precos_ipr2, precos_vbr, precos_rzn, tempo_ex_ipr, tempo_ex_vbr, tempo_ex_rzn
    global precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop, tempo_ex_ipr_postos
    global precos_vbr_janjao, tempo_ex_vbr_janjao
    global ipiranga_deu_erro, vibra_deu_erro, raizen_deu_erro, ipiranga_postos_deu_erro, vibra_janjao_deu_erro

    if nome_posto == "Ipiranga":
        try: 
            precos_ipr1, precos_ipr2, tempo_ex_ipr = preco_func()
            ipiranga_deu_erro = False
        except: 
            ipiranga_deu_erro = True     
            print("Erro ao coletar preços da Ipiranga")

    elif nome_posto == "Vibra":
        try: 
            precos_vbr, tempo_ex_vbr = preco_func()
            vibra_deu_erro = False
        except:
            vibra_deu_erro = True
            print("Erro ao coletar preços da Vibra")

    elif nome_posto == "Raizen":
        try: 
            precos_rzn, tempo_ex_rzn = preco_func()
            raizen_deu_erro = False
        except:
            raizen_deu_erro = True
            print("Erro ao coletar preços da Raízen")

    elif nome_posto == "Ipiranga Postos":
        try: 
            precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop, tempo_ex_ipr_postos = preco_func()
            ipiranga_postos_deu_erro = False
        except:
            ipiranga_postos_deu_erro = True
            print("Erro ao coletar preços da Ipiranga (Postos)")

    elif nome_posto == "Vibra Janjão":
        try: 
            precos_vbr_janjao, tempo_ex_vbr_janjao = preco_func()
            vibra_janjao_deu_erro = False
        except:
            vibra_janjao_deu_erro = True
            print("Erro ao coletar preços da Vibra (Janjão)")


inicio = time()

# Criando as threads para coletar os preços em paralelo
thread_ipr = threading.Thread(target=coleta_e_imprime, args=(ipiranga.coleta_precos, "Ipiranga"))
thread_vbr = threading.Thread(target=coleta_e_imprime, args=(vibra.coleta_precos, "Vibra"))
thread_rzn = threading.Thread(target=coleta_e_imprime, args=(raizen.coleta_precos, "Raizen"))
thread_ipr_postos = threading.Thread(target=coleta_e_imprime, args=(ipiranga_postos.coleta_precos, "Ipiranga Postos"))
thread_vbr_janjao = threading.Thread(target=coleta_e_imprime, args=(vibra_janjao.coleta_precos, "Vibra Janjão"))

# Iniciando as threads
thread_ipr.start()
thread_vbr.start()
thread_rzn.start()
thread_ipr_postos.start()
thread_vbr_janjao.start()

# Aguardando as threads terminarem
thread_ipr.join()
thread_vbr.join()
thread_rzn.join()
thread_ipr_postos.join()
thread_vbr_janjao.join()


# Tratando erros
combustiveis = ['cif_etanol', 'cif_gasolina', 'cif_gasolina_adt', 'cif_s10', 'cif_s500', 'fob_etanol', 'fob_gasolina', 'fob_gasolina_adt', 'fob_s10', 'fob_s500']

if ipiranga_deu_erro:
    precos_ipr1, precos_ipr2 = {}, {}
    for combustivel in combustiveis:
        precos_ipr1[combustivel] = None
        precos_ipr2[combustivel] = None
    tempo_ex_ipr = None

if vibra_deu_erro:
    precos_vbr = {}
    for combustivel in combustiveis:
        precos_vbr[combustivel] = None
    tempo_ex_vbr = None

if raizen_deu_erro:
    precos_rzn = {}
    for combustivel in combustiveis:
        precos_rzn[combustivel] = None
    tempo_ex_rzn = None

if ipiranga_postos_deu_erro:
    precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop = {}, {}, {}, {}, {}
    for combustivel in combustiveis:
        precos_gasstation[combustivel] = None
        precos_distrito[combustivel] = None
        precos_itirapua[combustivel] = None
        precos_ppp[combustivel] = None
        precos_pitstop[combustivel] = None
    tempo_ex_ipr_postos = None

if vibra_janjao_deu_erro:
    precos_vbr_janjao = {}
    for combustivel in combustiveis:
        precos_vbr_janjao[combustivel] = None
    tempo_ex_vbr_janjao = None


imprime_precos(
    precos_ipr1, precos_ipr2, tempo_ex_ipr,
    precos_vbr, tempo_ex_vbr,
    precos_rzn, tempo_ex_rzn,
    precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop, tempo_ex_ipr_postos,
    precos_vbr_janjao, tempo_ex_vbr_janjao,
    inicio
)

# Criando o DataFrame com os resultados
dados = cria_dicionario_para_df(
    precos_ipr1, precos_ipr2, precos_vbr, precos_rzn,
    precos_gasstation, precos_distrito, precos_itirapua, precos_ppp, precos_pitstop, tempo_ex_ipr_postos,
    precos_vbr_janjao,
)

df = pd.DataFrame(dados)

arquivo = f"C:/Users/titrr/OneDrive - MARIO ROBERTO TRANSP REVENDEDORA D OLEO DIESEL/Leva Diesel/Logística/resultados_precos.xlsx"

# Carregando o arquivo Excel existente (ou criando um novo se não existir)
try:
    writer = pd.ExcelWriter(arquivo, engine='openpyxl', datetime_format='dd/mm/yyyy HH:MM:SS', date_format='dd/mm/yyyy', mode='a', if_sheet_exists='replace')
except FileNotFoundError:
    writer = pd.ExcelWriter(arquivo, engine='openpyxl', datetime_format='dd/mm/yyyy HH:MM:SS', date_format='dd/mm/yyyy', mode='w')

df.to_excel(writer, index=False)
writer.close()

print(f"Resultados salvos em: {arquivo}")

# Abrindo o arquivo Excel com openpyxl
workbook = load_workbook(arquivo)
sheet = workbook['Sheet1']

# Inserindo uma linha acima de tudo
sheet.insert_rows(1)

# Mesclando células
sheet.merge_cells('B1:C1')
sheet.merge_cells('D1:E1')
sheet.merge_cells('F1:G1')
sheet.merge_cells('H1:I1')
sheet.merge_cells('J1:K1')
sheet.merge_cells('L1:M1')
sheet.merge_cells('N1:O1')
sheet.merge_cells('P1:Q1')
sheet.merge_cells('R1:S1')
sheet.merge_cells('T1:U1')

# Definindo o texto
sheet['B1'] = 'Pitstop'
sheet['D1'] = 'Distrito'
sheet['F1'] = 'Gas Station'
sheet['H1'] = 'Itirapuã'
sheet['J1'] = 'PPP'
sheet['L1'] = 'Janjão'
sheet['N1'] = 'TRR Ipiranga1'
sheet['P1'] = 'TRR Ipiranga2'
sheet['R1'] = 'TRR Vibra'
sheet['T1'] = 'TRR Raízen'

# Estilo para tornar o texto negrito
font_bold = Font(bold=True)
sheet['B1'].font = font_bold
sheet['D1'].font = font_bold
sheet['F1'].font = font_bold
sheet['H1'].font = font_bold
sheet['J1'].font = font_bold
sheet['L1'].font = font_bold
sheet['N1'].font = font_bold
sheet['P1'].font = font_bold
sheet['R1'].font = font_bold
sheet['T1'].font = font_bold

# Estilo para pintar o fundo da célula B1
amarelo = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
verde = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
roxo = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
sheet['B1'].fill = amarelo
sheet['D1'].fill = amarelo
sheet['F1'].fill = amarelo
sheet['H1'].fill = amarelo
sheet['J1'].fill = amarelo
sheet['L1'].fill = verde
sheet['N1'].fill = amarelo
sheet['P1'].fill = amarelo
sheet['R1'].fill = verde
sheet['T1'].fill = roxo

# Centralizando o texto em B1
centralizado = Alignment(horizontal='center', vertical='center')
sheet['B1'].alignment = centralizado
sheet['D1'].alignment = centralizado
sheet['F1'].alignment = centralizado
sheet['H1'].alignment = centralizado
sheet['J1'].alignment = centralizado
sheet['L1'].alignment = centralizado
sheet['N1'].alignment = centralizado
sheet['P1'].alignment = centralizado
sheet['R1'].alignment = centralizado
sheet['T1'].alignment = centralizado

# Adicionando bordas à célula B1
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)
sheet['B1'].border = border
sheet['D1'].border = border
sheet['F1'].border = border
sheet['H1'].border = border
sheet['J1'].border = border
sheet['L1'].border = border
sheet['N1'].border = border
sheet['P1'].border = border
sheet['R1'].border = border
sheet['T1'].border = border
sheet['U1'].border = border

# Salvando as alterações
workbook.save(arquivo)
print("Excel tratado com sucesso!")
